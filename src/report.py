"""
Excel output.

The report is the deliverable. Whoever picks this up wants to open one file,
see the headline numbers, and drill into whatever did not match. Everything
below serves that.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
INK = "1F2933"
ACCENT = "1F4E79"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BAND_FILL = PatternFill("solid", fgColor="F2F5F8")
FLAG_FILL = PatternFill("solid", fgColor="FDECEA")
MONEY = '#,##0.00;(#,##0.00);"-"'
DATE_FMT = "yyyy-mm-dd"
THIN = Side(style="thin", color="D6DBE0")
BOX = Border(bottom=THIN)


def _header(ws, row, headers):
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_table(ws, df, headers, widths, money_cols=(), date_cols=(), flag_col=None):
    _header(ws, 1, headers)
    _widths(ws, widths)
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BOX
            if c in money_cols:
                cell.number_format = MONEY
                cell.alignment = Alignment(horizontal="right")
            elif c in date_cols:
                cell.number_format = DATE_FMT
                cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0:
                cell.fill = BAND_FILL
        if flag_col:
            ws.cell(row=r, column=flag_col).fill = FLAG_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(df) + 1, 2)}"


def _title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=ACCENT)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, size=10, italic=True, color="6B7680")


def build(path, stats, matched, tol, un_gw, un_bk, layer_counts, reason_counts,
          batches=None, reversals=None, duplicates=None, review=None):
    wb = Workbook()

    # ---------------- Summary --------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    _widths(ws, [34, 18, 16, 4, 30, 14])
    _title(ws, "Transaction Reconciliation Report",
           f"Gateway export vs bank statement  |  {stats['period']}  |  "
           f"{stats['n_gw']:,} records matched in {stats['runtime']:.1f}s")

    rows = [
        ("Gateway records", stats["n_gw"], ""),
        ("Bank records", stats["n_bk"], ""),
        ("", "", ""),
        ("Matched one-to-one", "=COUNTA(Matched!A2:A100000)+COUNTA('Tolerance Matched'!A2:A100000)", ""),
        ("Settled inside a batch", stats["batch_payments"], ""),
        ("Reversed or returned", stats["n_reversals"], ""),
        ("Duplicate exports retired", stats["n_duplicates"], ""),
        ("Total accounted for", "=B7+B8+B10", ""),
        ("Coverage of gateway records", "=B11/B4", "0.0%"),
        ("", "", ""),
        ("Held for manual review", stats["n_review"], ""),
        ("Unmatched - gateway side", "=COUNTA('Unmatched Gateway'!A2:A100000)", ""),
        ("Unmatched - bank side", "=COUNTA('Unmatched Bank'!A2:A100000)", ""),
        ("", "", ""),
        ("Gateway gross value", stats["gw_value"], MONEY),
        ("Bank net value", stats["bk_value"], MONEY),
        ("Fees absorbed in settlement", stats["fee_total"], MONEY),
    ]
    r = 4
    for label, value, fmt in rows:
        if label:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = Font(name=FONT, size=11, color=INK)
            vc = ws.cell(row=r, column=2, value=value)
            vc.font = Font(name=FONT, size=11, bold=True, color=ACCENT)
            vc.alignment = Alignment(horizontal="right")
            if fmt and fmt != MONEY:
                vc.number_format = fmt
            elif fmt == MONEY:
                vc.number_format = MONEY
            else:
                vc.number_format = "#,##0"
        r += 1

    ws["A22"] = "Matches by layer"
    ws["A22"].font = Font(name=FONT, size=11, bold=True, color=ACCENT)
    _header(ws, 23, ["Layer", "Count", "Share"])
    ws.freeze_panes = None
    rr = 24
    total_matched = sum(layer_counts.values())
    for layer, count in layer_counts.items():
        ws.cell(row=rr, column=1, value=layer).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=2, value=count).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=2).number_format = "#,##0"
        share = ws.cell(row=rr, column=3, value=f"=B{rr}/{total_matched}")
        share.number_format = "0.0%"
        share.font = Font(name=FONT, size=10, color=INK)
        rr += 1

    pie = PieChart()
    pie.title = "Matches by layer"
    pie.height, pie.width = 7, 10
    pie.add_data(Reference(ws, min_col=2, min_row=23, max_row=rr - 1), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=24, max_row=rr - 1))
    ws.add_chart(pie, "E4")

    start = rr + 2
    ws.cell(row=start, column=1, value="Unmatched items by reason").font = Font(
        name=FONT, size=11, bold=True, color=ACCENT)
    _header(ws, start + 1, ["Reason", "Count"])
    ws.freeze_panes = None
    rr = start + 2
    for reason, count in reason_counts.items():
        ws.cell(row=rr, column=1, value=reason).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=2, value=count).font = Font(name=FONT, size=10, color=INK)
        ws.cell(row=rr, column=2).number_format = "#,##0"
        rr += 1

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Unmatched by reason"
    bar.height, bar.width = 8, 12
    bar.legend = None
    bar.add_data(Reference(ws, min_col=2, min_row=start + 1, max_row=rr - 1),
                 titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start + 2, max_row=rr - 1))
    ws.add_chart(bar, "E20")

    note = ws.cell(row=rr + 2, column=1,
                   value="Built with synthetic data. No client information appears "
                         "in this workbook.")
    note.font = Font(name=FONT, size=9, italic=True, color="8A94A0")

    # ---------------- Matched --------------------------------------------
    ws = wb.create_sheet("Matched")
    _write_table(
        ws, matched,
        ["Transaction ID", "Gateway Date", "Gateway Amount", "Bank Reference",
         "Bank Date", "Bank Amount", "Difference", "Confidence"],
        [22, 14, 16, 24, 12, 14, 12, 12],
        money_cols=(3, 6, 7), date_cols=(2, 5),
    )

    # ---------------- Tolerance ------------------------------------------
    ws = wb.create_sheet("Tolerance Matched")
    _write_table(
        ws, tol,
        ["Transaction ID", "Gateway Date", "Gateway Amount", "Bank Reference",
         "Bank Date", "Bank Amount", "Difference", "Match Basis", "Confidence"],
        [22, 14, 16, 24, 12, 14, 12, 34, 12],
        money_cols=(3, 6, 7), date_cols=(2, 5),
    )

    # ---------------- Unmatched ------------------------------------------
    ws = wb.create_sheet("Unmatched Gateway")
    _write_table(
        ws, un_gw,
        ["Transaction ID", "Date", "Amount", "Channel", "Expected Settlement",
         "Status", "Likely Reason"],
        [22, 20, 14, 10, 18, 20, 56],
        money_cols=(3,), date_cols=(2, 5), flag_col=7,
    )

    ws = wb.create_sheet("Unmatched Bank")
    _write_table(
        ws, un_bk,
        ["Reference", "Date", "Amount", "Description", "Likely Reason"],
        [24, 14, 14, 40, 44],
        money_cols=(3,), date_cols=(2,), flag_col=5,
    )

    # ---------------- Batch settlements ----------------------------------
    if batches is not None and len(batches):
        ws = wb.create_sheet("Batch Settlements")
        _write_table(
            ws, batches,
            ["Batch ID", "Bank Reference", "Settlement Date", "Payments Matched",
             "Count Declared", "Gross Value", "Net Credited", "Aggregate Fees",
             "Effective Fee Rate", "Held Back", "Reconciles"],
            [14, 24, 16, 17, 15, 15, 15, 15, 16, 11, 40],
            money_cols=(6, 7, 8), date_cols=(3,), flag_col=11,
        )
        for r in range(2, len(batches) + 2):
            ws.cell(row=r, column=9).number_format = "0.00%"

    # ---------------- Reversals and returns -------------------------------
    if reversals is not None and len(reversals):
        ws = wb.create_sheet("Reversals")
        _write_table(
            ws, reversals,
            ["Original Reference", "Original Date", "Original Amount",
             "Reversal Reference", "Reversal Date", "Reversal Amount",
             "Residual", "Return Code", "Type"],
            [24, 14, 16, 24, 14, 16, 12, 13, 12],
            money_cols=(3, 6, 7), date_cols=(2, 5),
        )

    # ---------------- Held for review -------------------------------------
    if review is not None and len(review):
        ws = wb.create_sheet("Needs Review")
        _write_table(
            ws, review,
            ["Transaction ID", "Date", "Amount", "Channel", "Candidates",
             "Top Candidates", "Why Held"],
            [22, 20, 14, 10, 12, 52, 56],
            money_cols=(3,), date_cols=(2,), flag_col=7,
        )

    # ---------------- Duplicates ------------------------------------------
    if duplicates is not None and len(duplicates):
        ws = wb.create_sheet("Duplicate Exports")
        _write_table(
            ws, duplicates,
            ["Transaction ID", "Date", "Amount", "Customer Ref",
             "Duplicate Of", "Basis"],
            [22, 20, 14, 14, 22, 46],
            money_cols=(3,), date_cols=(2,), flag_col=6,
        )

    wb.save(path)
